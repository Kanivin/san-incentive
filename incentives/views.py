from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import check_password
from datetime import datetime
from .models import UserProfile, Deal, Role, Segment, Module, AnnualTarget, TargetTransaction, Transaction, PayoutTransaction, IncentiveSetup, SetupChargeSlab, TopperMonthSlab, HighValueDealSlab, Permission
from .forms import UserProfileForm, SegmentForm,DealForm, AnnualTargetForm,  RoleForm, ModuleForm, IncentiveSetupForm, SetupChargeSlabForm, TopperMonthSlabForm, HighValueDealSlabForm
from django.forms import ValidationError, modelformset_factory
from decimal import Decimal
import json
import logging
from django.db.models import Count
from incentives.utils.incentive_engine import DealRuleEngine 
from django.core.paginator import Paginator
from django.contrib.auth.views import LogoutView
from django.template.loader import get_template
from django.urls import reverse_lazy

from incentives.utils.db_backup import upload_db_to_gcs
from django.http import HttpResponse
from django.http import JsonResponse
from .models import JobRunLog, ScheduledJob, ChangeLog
from .tasks import monthly_sales_incentive, annual_target_achievement
from django.db.models import Sum
from collections import defaultdict
from django.db.models.functions import ExtractMonth
from django.utils import timezone
import tablib
import os
from django.conf import settings

class CustomLogoutView(LogoutView):
    next_page = reverse_lazy('login') 

logger = logging.getLogger(__name__)
# ---------- Authentication Views ----------
def validate_form_data_length(*args):
    """
    Validates that all input lists are of the same non-zero length.
    Returns True if valid, False otherwise.
    """
    if not args:
        return False
    length = len(args[0])
    return all(len(lst) == length and length > 0 for lst in args)

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('mail_id')
        password = request.POST.get('password')
        next_url = request.GET.get('next')

        try:
            user = UserProfile.objects.get(mail_id=username)
            if check_password(password, user.password):
                request.session['mail_id'] = user.mail_id
                request.session['user_id'] = user.id
                request.session['role_id'] = user.user_type.id
                user_type = user.user_type.name.lower()

                request.session['user_type'] = user_type
                if user_type == 'superadmin':
                    return redirect(next_url or 'dashboard_router')
                elif user_type == 'accounts':
                    return redirect('dashboard_router')
                elif user_type == 'admin':
                    return redirect('dashboard_router')
                else:
                    return redirect('dashboard_router')
            else:
                messages.error(request, 'Invalid username or password.')
        except UserProfile.DoesNotExist:
            messages.error(request, 'Invalid username or password.')

    return render(request, 'login.html')


# ---------- Dashboard Views ----------

def admin_dashboard(request):
    return render(request, 'owner/dashboard.html')

def accounts_dashboard(request):
    return render(request, 'accounts/dashboard.html')

def superadmin_dashboard(request):
    return render(request, 'super/dashboard.html')

def sales_dashboard(request):
    user_id = request.session.get('user_id')   
    user_type = request.session.get('user_type')
    current_year = datetime.now().year    
    financial_years = [current_year - i for i in range(5)]    

    if user_type == 'admin' or user_type == 'superadmin':

        try:
            current_profile = UserProfile.objects.get(id=user_id)
        except UserProfile.DoesNotExist:
            # Handle missing profile case gracefully
            return render(request, 'dashboard/error.html', {'message': 'User profile not found.'})
       
        team_members_qs = UserProfile.objects.filter(user_type__name__in=['saleshead', 'salesperson'])
        team_members = list(team_members_qs)

        selected_user_id = request.GET.get('team_member')
        selected_user = current_profile  # default
       
        if selected_user_id:
            try:
                selected_user = UserProfile.objects.get(id=selected_user_id)
            except UserProfile.DoesNotExist:
                selected_user = current_profile.id
        else:
            selected_user_id = current_profile.id

        
        selected_year = request.GET.get('financial_year')

        if selected_year:
            try:
                selected_year = int(selected_year)
            except ValueError:
                selected_year = current_year
        else:
            selected_year = current_year

        payouts = PayoutTransaction.objects.filter(user=selected_user_id,payout_date__year=selected_year)
       
        paid_payouts = payouts.filter(payout_status='Paid')
        pending_payouts = payouts.filter(payout_status='Pending')
        readytopay_payouts = payouts.filter(payout_status='ReadyToPay')

        total_payout = paid_payouts.aggregate(total=Sum('payout_amount'))['total'] or 0
        total_pending_payout = pending_payouts.aggregate(total=Sum('payout_amount'))['total'] or 0
        total_readytopay_payout = readytopay_payouts.aggregate(total=Sum('payout_amount'))['total'] or 0
        
        # 1. Total Earned Target Amount this year
        total_target = TargetTransaction.objects.filter(
            user_id=selected_user_id,
            eligibility_status='Eligible',
            deal__dealWonDate__year=selected_year
        ).aggregate(total=Sum('amount'))['total'] or Decimal(0)

        print(f"Warning: total_target is {total_target} .")  
        overall_total_target = TargetTransaction.objects.filter(
                    user_id=selected_user_id,
                    eligibility_status__in=['Ineligible', 'Eligible'],
                    deal__dealWonDate__year=selected_year
        ).aggregate(total=Sum('amount'))['total'] or Decimal(0)
            
        print(f"Warning: overall_total_target is {overall_total_target} .")      
        # 2. Get Annual Target
        try:
            annual_target = AnnualTarget.objects.get(employee_id=selected_user_id, financial_year=selected_year)
            annual_target_amount = annual_target.annual_target_amount
            net_salary = annual_target.net_salary
        except AnnualTarget.DoesNotExist:
            annual_target_amount = Decimal(0)
            net_salary = Decimal(0)

        # 3. Compute Target % Achievement
        if annual_target_amount > 0:
            target_percentage = (total_target / annual_target_amount) * 100
        else:
            target_percentage = Decimal(0)

        print(f"Warning: target_percentage is {target_percentage} .")  

        if overall_total_target > 0 and annual_target_amount > 0:
            overall_target_percentage = (overall_total_target / annual_target_amount) * 100
        else:
            overall_target_percentage = Decimal(0)    

        print(f"Warning: overall_target_percentage is {overall_target_percentage} .")  

        # 4. Annual Incentive Calculation (based on slabs)
        annual_target_incentive = Decimal(0)
        if target_percentage >= 100:
            annual_target_incentive = net_salary * Decimal('2.0')  # 200%
        elif target_percentage >= 95:
            annual_target_incentive = net_salary * Decimal('1.5')  # 150%
        elif target_percentage >= 90:
            annual_target_incentive = net_salary * Decimal('1.25')  # 125%
        elif target_percentage >= 75:
            annual_target_incentive = net_salary * Decimal('1.0')  # 100%
        else:
            annual_target_incentive = Decimal(0)  # Below 75% gets no incentive

        # 5. Subscription Incentive %
        setup = IncentiveSetup.objects.filter(financial_year=str(selected_year)).order_by('-created_at').first()

        if target_percentage >= 100:
            subscription_incentive_percent = setup.subscription_100_per_target or Decimal('0.00')
        elif target_percentage >= 75:
            subscription_incentive_percent = setup.subscription_75_per_target or Decimal('0.00')
        elif target_percentage >= 50:
            subscription_incentive_percent = setup.subscription_50_per_target or Decimal('0.00')
        elif target_percentage < 50:
            #subscription_incentive_percent = setup.subscription_below_50_per or Decimal('0.00')
            subscription_incentive_percent = Decimal('0.00')
        else: 
            subscription_incentive_percent = Decimal('0.00')

        # 6. Subscription Incentive = % of target achieved amount
        subscription_incentive = total_target * (subscription_incentive_percent / 100)

        # --- Product-Wise Payout Calculation ---
        product_wise_data = PayoutTransaction.objects.filter(user_id=selected_user_id, payout_date__year=selected_year).values('deal__segment__name', 'incentive_person_type').annotate(total=Sum('payout_amount')).order_by('-total')
        product_wise_labels = []
        product_wise_series = []
        product_wise_colors = ['#008FFB', '#00E396', '#FEB019', '#FF4560', '#775DD0', '#3F51B5', '#546E7A', '#546E7A', '#26A69A','#D10CE8','#F86624','#2E93fA','#66DA26']
        
        segment_names = list(Segment.objects.values_list('name', flat=True).distinct())

        group_map = defaultdict(float)

        for i, row in enumerate(product_wise_data):
            segment_name = row['deal__segment__name']
            fallback = row['incentive_person_type']
            if segment_name:
                group_name = segment_name
            else:
                # Try to match any known segment name inside incentive_person_type
                match = next((s for s in segment_names if s.lower() in fallback.lower()), None)
                group_name = match if match else (fallback or f"Uncategorized {i+1}")

            group_map[group_name] += float(row['total'] or 0)
            
        product_wise_labels = list(group_map.keys())
        product_wise_series = list(group_map.values())
        product_wise_colors = product_wise_colors[:len(product_wise_labels)]

        monthly_achievement = [0] * 12
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
       # Step 1: Get per-month totals
        monthly_data = TargetTransaction.objects.filter(
            user_id=selected_user_id,
            eligibility_status__in=['Ineligible', 'Eligible'],
            deal__dealWonDate__year=selected_year
        ).annotate(
            month=ExtractMonth('deal__dealWonDate')
        ).values('month').annotate(total=Sum('amount'))

        # Step 2: Fill raw monthly totals
        raw_monthly_totals = [0] * 12
        for row in monthly_data:
            month_idx = row['month'] - 1
            raw_monthly_totals[month_idx] = float(row['total'] or 0)

        # Step 3: Convert to cumulative
        cumulative_total = 0
        for i in range(12):
            cumulative_total += raw_monthly_totals[i]
            monthly_achievement[i] = round(cumulative_total, 2)
                
        context = {
            'team_members': team_members,
            'selected_user': selected_user,
            'total_payout': total_payout,
            'total_pending_payout': total_pending_payout,
            'total_readytopay_payout': total_readytopay_payout,            
            'annual_target_incentive': annual_target_incentive,
            'subscription_incentive': subscription_incentive,
            'target_percentage': round(target_percentage, 2),
            'overall_target_percentage': round(overall_target_percentage, 2),            
            'financial_years': financial_years,
            'selected_year': selected_year,
            'paid_payouts': paid_payouts,
            'readytopay_payouts': readytopay_payouts,            
            'pending_payouts': pending_payouts,
            'product_wise_series': json.dumps(product_wise_series),
            'product_wise_labels': json.dumps(product_wise_labels),
            'product_wise_colors': json.dumps(product_wise_colors),
            'months_json': json.dumps(months),
            'achievement_json': json.dumps(monthly_achievement),
            'total_target': float(annual_target_amount),
            'gross_monthly_salary': float(net_salary),
        }
        return render(request, 'sales/dashboard.html', context)

def dashboard_router(request):    
    user_id = request.session.get('user_id')   
    user_type = request.session.get('user_type')
    current_year = datetime.now().year    
    financial_years = [current_year - i for i in range(5)]    

    if user_type == 'saleshead' or user_type == 'salesperson':

        try:
            current_profile = UserProfile.objects.get(id=user_id)
        except UserProfile.DoesNotExist:
            # Handle missing profile case gracefully
            return render(request, 'dashboard/error.html', {'message': 'User profile not found.'})
       
        team_members_qs = UserProfile.objects.filter(team_head=user_id)
        # Users reporting to the current user (team members)
        current_profile.fullname += " - Self"
        team_members = [current_profile] + list(team_members_qs)

        selected_user_id = request.GET.get('team_member')
        selected_user = current_profile  # default
       
        if selected_user_id:
            try:
                selected_user = UserProfile.objects.get(id=selected_user_id)
            except UserProfile.DoesNotExist:
                selected_user = current_profile
        else:
            selected_user_id = current_profile.id

        
        selected_year = request.GET.get('financial_year')

        if selected_year:
            try:
                selected_year = int(selected_year)
            except ValueError:
                selected_year = current_year
        else:
            selected_year = current_year

        payouts = PayoutTransaction.objects.filter(user=selected_user_id,payout_date__year=selected_year)
       
        paid_payouts = payouts.filter(payout_status='Paid')
        pending_payouts = payouts.filter(payout_status='Pending')
        readytopay_payouts = payouts.filter(payout_status='ReadyToPay')

        total_payout = paid_payouts.aggregate(total=Sum('payout_amount'))['total'] or 0
        total_pending_payout = pending_payouts.aggregate(total=Sum('payout_amount'))['total'] or 0
        total_readytopay_payout = readytopay_payouts.aggregate(total=Sum('payout_amount'))['total'] or 0
        
        # 1. Total Earned Target Amount this year
        total_target = TargetTransaction.objects.filter(
            user_id=selected_user_id,
            eligibility_status='Eligible',
            deal__dealWonDate__year=selected_year
        ).aggregate(total=Sum('amount'))['total'] or Decimal(0)

        print(f"Warning: total_target is {total_target} .")  
        overall_total_target = TargetTransaction.objects.filter(
                    user_id=selected_user_id,
                    eligibility_status__in=['Ineligible', 'Eligible'],
                    deal__dealWonDate__year=selected_year
        ).aggregate(total=Sum('amount'))['total'] or Decimal(0)
            
        print(f"Warning: overall_total_target is {overall_total_target} .")      
        # 2. Get Annual Target
        try:
            annual_target = AnnualTarget.objects.get(employee_id=selected_user_id, financial_year=selected_year)
            annual_target_amount = annual_target.annual_target_amount
            net_salary = annual_target.net_salary
        except AnnualTarget.DoesNotExist:
            annual_target_amount = Decimal(0)
            net_salary = Decimal(0)

        # 3. Compute Target % Achievement
        if annual_target_amount > 0:
            target_percentage = (total_target / annual_target_amount) * 100
        else:
            target_percentage = Decimal(0)

        print(f"Warning: target_percentage is {target_percentage} .")  

        if overall_total_target > 0 and annual_target_amount > 0:
            overall_target_percentage = (overall_total_target / annual_target_amount) * 100
        else:
            overall_target_percentage = Decimal(0)    

        print(f"Warning: overall_target_percentage is {overall_target_percentage} .")  

        # 4. Annual Incentive Calculation (based on slabs)
        annual_target_incentive = Decimal(0)
        if target_percentage >= 100:
            annual_target_incentive = net_salary * Decimal('2.0')  # 200%
        elif target_percentage >= 95:
            annual_target_incentive = net_salary * Decimal('1.5')  # 150%
        elif target_percentage >= 90:
            annual_target_incentive = net_salary * Decimal('1.25')  # 125%
        elif target_percentage >= 75:
            annual_target_incentive = net_salary * Decimal('1.0')  # 100%
        else:
            annual_target_incentive = Decimal(0)  # Below 75% gets no incentive

        # 5. Subscription Incentive %
        setup = IncentiveSetup.objects.filter(financial_year=str(selected_year)).order_by('-created_at').first()

        if target_percentage >= 100:
            subscription_incentive_percent = setup.subscription_100_per_target or Decimal('0.00')
        elif target_percentage >= 75:
            subscription_incentive_percent = setup.subscription_75_per_target or Decimal('0.00')
        elif target_percentage >= 50:
            subscription_incentive_percent = setup.subscription_50_per_target or Decimal('0.00')
        elif target_percentage < 50:
            #subscription_incentive_percent = setup.subscription_below_50_per or Decimal('0.00')
            subscription_incentive_percent = Decimal('0.00')
        else: 
            subscription_incentive_percent = Decimal('0.00')

        # 6. Subscription Incentive = % of target achieved amount
        subscription_incentive = total_target * (subscription_incentive_percent / 100)

        # --- Product-Wise Payout Calculation ---
        product_wise_data = PayoutTransaction.objects.filter(user_id=selected_user_id, payout_date__year=selected_year).values('deal__segment__name', 'incentive_person_type').annotate(total=Sum('payout_amount')).order_by('-total')
        product_wise_labels = []
        product_wise_series = []
        product_wise_colors = ['#008FFB', '#00E396', '#FEB019', '#FF4560', '#775DD0', '#3F51B5', '#546E7A', '#546E7A', '#26A69A','#D10CE8','#F86624','#2E93fA','#66DA26']
        
        segment_names = list(Segment.objects.values_list('name', flat=True).distinct())

        group_map = defaultdict(float)

        for i, row in enumerate(product_wise_data):
            segment_name = row['deal__segment__name']
            fallback = row['incentive_person_type']
            if segment_name:
                group_name = segment_name
            else:
                # Try to match any known segment name inside incentive_person_type
                match = next((s for s in segment_names if s.lower() in fallback.lower()), None)
                group_name = match if match else (fallback or f"Uncategorized {i+1}")

            group_map[group_name] += float(row['total'] or 0)
            
        product_wise_labels = list(group_map.keys())
        product_wise_series = list(group_map.values())
        product_wise_colors = product_wise_colors[:len(product_wise_labels)]

        monthly_achievement = [0] * 12
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
       # Step 1: Get per-month totals
        monthly_data = TargetTransaction.objects.filter(
            user_id=selected_user_id,
            eligibility_status__in=['Ineligible', 'Eligible'],
            deal__dealWonDate__year=selected_year
        ).annotate(
            month=ExtractMonth('deal__dealWonDate')
        ).values('month').annotate(total=Sum('amount'))

        # Step 2: Fill raw monthly totals
        raw_monthly_totals = [0] * 12
        for row in monthly_data:
            month_idx = row['month'] - 1
            raw_monthly_totals[month_idx] = float(row['total'] or 0)

        # Step 3: Convert to cumulative
        cumulative_total = 0
        for i in range(12):
            cumulative_total += raw_monthly_totals[i]
            monthly_achievement[i] = round(cumulative_total, 2)
                
        context = {
            'team_members': team_members,
            'selected_user': selected_user,
            'total_payout': total_payout,
            'total_pending_payout': total_pending_payout,
            'total_readytopay_payout': total_readytopay_payout,            
            'annual_target_incentive': annual_target_incentive,
            'subscription_incentive': subscription_incentive,
            'target_percentage': round(target_percentage, 2),
            'overall_target_percentage': round(overall_target_percentage, 2),            
            'financial_years': financial_years,
            'selected_year': selected_year,
            'paid_payouts': paid_payouts,
            'readytopay_payouts': readytopay_payouts,            
            'pending_payouts': pending_payouts,
            'product_wise_series': json.dumps(product_wise_series),
            'product_wise_labels': json.dumps(product_wise_labels),
            'product_wise_colors': json.dumps(product_wise_colors),
            'months_json': json.dumps(months),
            'achievement_json': json.dumps(monthly_achievement),
            'total_target': float(annual_target_amount),
            'gross_monthly_salary': float(net_salary),
        }
        return render(request, 'sales/dashboard.html', context)
       
    else:
        context = {}

        selected_year = request.GET.get('financial_year')

        if selected_year:
            try:
                selected_year = int(selected_year)
            except ValueError:
                selected_year = current_year
        else:
            selected_year = current_year
        previous_year = selected_year - 1

        context['selected_year'] = selected_year
        context['financial_years'] = financial_years
        # Dynamic Counts
        context['employee_count'] = UserProfile.objects.count()
        context['client_count'] = Deal.objects.values('clientName').distinct().count()
        context['product_count'] = Segment.objects.values('name').distinct().count()

        # Total Payout
        context['total_payout'] = (
            PayoutTransaction.objects.filter(payout_status='Paid')
            .aggregate(total=Sum('payout_amount'))['total'] or Decimal(0)
        )

        # Top Performers (Highest Target Achievement %)
        
        top_performers = []

        targets = AnnualTarget.objects.select_related('employee').filter(financial_year=str(selected_year))
        for target in targets:
            earned = (
                TargetTransaction.objects.filter(
                    user=target.employee,
                    eligibility_status='Eligible',
                    deal__dealWonDate__year=selected_year
                )
                .aggregate(total=Sum('amount'))['total'] or Decimal(0)
            )

            achievement = float((earned / target.annual_target_amount) * 100) if target.annual_target_amount > 0 else 0
            top_performers.append({
                'name': target.employee.fullname,
                'email': target.employee.mail_id,
                'target': target.annual_target_amount,
                'actual': earned,
                'achievement' : achievement,
                'priority': 'High' if achievement > 90 else 'Medium' if achievement > 75 else 'Low',
            })

        context['top_performers'] = sorted(top_performers, key=lambda x: x['actual'], reverse=True)[:5]

        def get_monthly_payouts(year):
            monthly_totals = [0] * 12
            payouts = (
                PayoutTransaction.objects.filter(payout_status='Paid', payout_date__year=year)
                .annotate(month=ExtractMonth('payout_date'))
                .values('month')
                .annotate(total=Sum('payout_amount'))
            )
            for entry in payouts:
                monthly_totals[entry['month'] - 1] = float(entry['total']) if entry['total'] else 0
            return monthly_totals


        context['earnings_incentives_data'] = json.dumps({
            'current_year': selected_year,
            'previous_year': previous_year,
            'months': ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'],
            'series': [
                {
                    'name': 'Current Year',
                    'data': get_monthly_payouts(selected_year)
                },
                {
                    'name': 'Previous Year',
                    'data': get_monthly_payouts(previous_year)
                },
            ]
        })
        
        return render(request, 'owner/dashboard.html', context)

# ---------- User Management Views ----------
def user_list(request):
    search_query = request.GET.get('search', '').strip()
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()

    users = UserProfile.objects.all().order_by('fullname')

    if search_query:
        users = users.filter(
            Q(fullname__icontains=search_query) |
            Q(employee_id__icontains=search_query) |
            Q(mail_id__icontains=search_query)
        )

    if start_date:
        users = users.filter(created_at__gte=start_date)
    if end_date:
        users = users.filter(created_at__lte=end_date)

    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(users, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'owner/users/user_list.html', context)


def export_users_xlsx(request):
    # Same filters as in user_list
    search_query = request.GET.get('search', '').strip()
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()

    users = UserProfile.objects.all().order_by('fullname')

    if search_query:
        users = users.filter(
            Q(fullname__icontains=search_query) |
            Q(employee_id__icontains=search_query) |
            Q(mail_id__icontains=search_query)
        )

    if start_date:
        users = users.filter(created_at__gte=start_date)
    if end_date:
        users = users.filter(created_at__lte=end_date)

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    # Header row
    headers = ['Employee ID', 'Full Name', 'Mail ID', 'User Type', 'Created At']
    ws.append(headers)

    # Data rows
    for user in users:
        ws.append([
            user.employee_id,
            user.fullname,
            user.mail_id,
            user.user_type.name if user.user_type else '',
            user.created_at.strftime('%Y-%m-%d') if user.created_at else '',
        ])

    # Prepare HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
   
    response['Content-Disposition'] = f'attachment; filename="users.xlsx"'

    wb.save(response)
    return response


def export_users_pdf(request):
    search_query = request.GET.get('search', '').strip()
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()


    users = UserProfile.objects.all().order_by('fullname')

    if search_query:
        users = users.filter(
            Q(fullname__icontains=search_query) |
            Q(employee_id__icontains=search_query) |
            Q(mail_id__icontains=search_query)
        )

    if start_date:
        users = users.filter(created_at__gte=start_date)
    if end_date:
        users = users.filter(created_at__lte=end_date)


    template = get_template('owner/users/user_pdf.html')  # Create this template
    html = template.render({'users': users})
   
    # Create the PDF object
    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = f'attachment; filename="users.pdf"'

    pisa.CreatePDF(html, dest=response)
    return response


def user_create(request):
    if request.method == 'POST':
        form = UserProfileForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            # If you have a password field, don't forget:
            # user.set_password(form.cleaned_data['password'])
            user.save()
            return redirect('user_list')
        else:
            # 🔥 Print the form errors in console
            print("Form errors:", form.errors)
    else:
        form = UserProfileForm()
    
    roles = Role.objects.exclude(name='superadmin')
    users = UserProfile.objects.filter(user_type__name__in=['admin', 'saleshead'])
    return render(request, 'owner/users/user_form.html', {
        'form': form,
        'roles': roles,
        'users':users,
        'title': 'Create User',
    })

def user_edit(request, pk):
    user = get_object_or_404(UserProfile, pk=pk)
    roles = Role.objects.all()
    users = UserProfile.objects.filter(user_type__name__in=['admin', 'saleshead'])
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            return redirect('user_list')
    else:
        form = UserProfileForm(instance=user)

    return render(request, 'owner/users/user_form.html', {
        'form': form,
        'roles': roles,
        'users':users,
        'title': 'Edit User'
    })

def user_delete(request, pk):
    user = get_object_or_404(UserProfile, pk=pk)
    if request.method == 'POST':
        user.delete()
        return redirect('user_list')
    return render(request, 'owner/users/user_confirm_delete.html', {
        'user': user,
        'title': 'Delete User'
    })


# ---------- Deal Management Views ----------
from django.shortcuts import render, redirect, get_object_or_404
from .forms import DealForm
from .models import Deal, UserProfile
from pprint import pprint
def deal_list(request):
    search_query = request.GET.get('search', '')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    deals = Deal.objects.all()

    # Filter by search
    if search_query:
        deals = deals.filter(
            Q(clientName__icontains=search_query) |
            Q(segment__name__icontains=search_query) |
            Q(dealType__icontains=search_query)
        )

    # Filter by date range
    if start_date and end_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            end = datetime.strptime(end_date, "%Y-%m-%d")
            deals = deals.filter(dealWonDate__range=(start, end))
        except ValueError:
            pass

    paginator = Paginator(deals.order_by('-dealWonDate'), 10)  # 10 per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'owner/deal/deal_list.html', {
        'page_obj': page_obj,
        'search_query': search_query,
        'start_date': start_date,
        'end_date': end_date,
    })
def deal_export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deals"

    # Header row
    ws.append(['S.No', 'Client Name', 'Segment', 'Deal Type', 'Deal Won Date', 'Status'])

    # Data rows
    deals = Deal.objects.all()
    for i, deal in enumerate(deals, start=1):
        ws.append([
            i,
            deal.clientName,
            str(deal.segment),      # convert FK to string (e.g. segment.name or __str__)
            deal.dealType,
            deal.dealWonDate.strftime('%Y-%m-%d') if deal.dealWonDate else '',
            deal.status
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="deals.xlsx"'
    wb.save(response)
    return response

def deal_export_pdf(request):
    deals = Deal.objects.all()
    template = get_template('owner/deal/deal_pdf.html')  # Create this template
    html = template.render({'deals': deals})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="deals.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response

def upload_to_gcs(uploaded_file, bucket_name, destination_blob_name):
    from google.cloud import storage
    client = storage.Client()
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(destination_blob_name)
    blob.upload_from_file(uploaded_file, rewind=True)

def upload_to_locally(uploaded_file, bucket_name, destination_blob_name):
    base_dir = settings.MEDIA_ROOT
    full_path = os.path.join(base_dir, destination_blob_name)
    # Create target directory if it doesn't exist
    os.makedirs(os.path.dirname(full_path), exist_ok=True)
    # Write file
    with open(full_path, 'wb+') as f:
        for chunk in uploaded_file.chunks():
            f.write(chunk)


def deal_create(request):
    # Fetch users with user_type 'salesperson' or 'saleshead'
    users = UserProfile.objects.filter(user_type__name__in=['salesperson', 'saleshead'])

    user_id = request.session.get('user_id') 
     
    try:
        # This line (user_profile = UserProfile.objects.get(id=user_id))
        # and subsequent lines within the 'try' block MUST be indented.
        user_profile = UserProfile.objects.get(id=user_id)
        isCoordinator = user_profile.coordinator
    except UserProfile.DoesNotExist:
        # All lines within the 'except' block MUST be indented.
        isCoordinator = False
        print(f"Warning: UserProfile with ID {user_id} does not exist.")
    except Exception as e:
        # All lines within this 'except' block MUST be indented.
        isCoordinator = False
        print(f"An unexpected error occurred: {e}")    

    if request.method == 'POST':
        form = DealForm(request.POST)
        import uuid, os
        if request.FILES.get('refDocs'):
            uploaded_file = request.FILES['refDocs']
            if uploaded_file.size > 10 * 1024 * 1024:
                form.add_error('refDocs', 'File too large (max 10MB)')
            else:
                base_name = os.path.splitext(uploaded_file.name)[0]
                ext = os.path.splitext(uploaded_file.name)[1]
                unique_id = uuid.uuid4().hex[:8]
                safe_file_name = f"{base_name}_{unique_id}{ext}"
                gcs_path = f"deals/ref_docs/{safe_file_name}"
                #upload_to_gcs(uploaded_file, 'san-incentive', gcs_path)
                upload_to_locally(uploaded_file, 'san-incentive', gcs_path)
                           

        if form.is_valid():
            deal = form.save(commit=False)
            if request.FILES.get('refDocs'):
                deal.refDocs.name = gcs_path         
            deal.created_by=request.session.get('mail_id')
            deal.save()
            return redirect('deal_list')
    else:
        form = DealForm()

    return render(request, 'owner/deal/deal_form.html', {
        'form': form,
        'action': 'Create',
        'isCoordinator': isCoordinator,
        'title': 'Create Deal',
        'users': users,
    })

def deal_view(request, pk):
    deal = get_object_or_404(Deal, pk=pk)

    # Fetch users who are either salesperson or saleshead (if needed for display)
    users = UserProfile.objects.filter(user_type__name__in=['salesperson', 'saleshead'])

    return render(request, 'owner/deal/deal_view.html', {
        'deal': deal,
        'users': users,
        'title': 'View Deal',
    })

def deal_update(request, pk):
    deal = get_object_or_404(Deal, pk=pk)

    # Fetch the users that are either salesperson or saleshead
    users = UserProfile.objects.filter(user_type__name__in=['salesperson', 'saleshead'])
    user_id = request.session.get('user_id') 
    
    try:
        # This line (user_profile = UserProfile.objects.get(id=user_id))
        # and subsequent lines within the 'try' block MUST be indented.
        user_profile = UserProfile.objects.get(id=user_id)
        isCoordinator = user_profile.coordinator
    except UserProfile.DoesNotExist:
        # All lines within the 'except' block MUST be indented.
        isCoordinator = False
        print(f"Warning: UserProfile with ID {user_id} does not exist.")
    except Exception as e:
        # All lines within this 'except' block MUST be indented.
        isCoordinator = False
        print(f"An unexpected error occurred: {e}") 

    if request.method == 'POST':
        form = DealForm(request.POST, instance=deal)

        import uuid, os
        if request.FILES.get('refDocs'):
            uploaded_file = request.FILES['refDocs']
            if uploaded_file.size > 10 * 1024 * 1024:
                form.add_error('refDocs', 'File too large (max 10MB)')
            else:
                base_name = os.path.splitext(uploaded_file.name)[0]
                ext = os.path.splitext(uploaded_file.name)[1]
                unique_id = uuid.uuid4().hex[:8]
                safe_file_name = f"{base_name}_{unique_id}{ext}"
                gcs_path = f"deals/ref_docs/{safe_file_name}"
                #upload_to_gcs(uploaded_file, 'san-incentive', gcs_path)
                upload_to_locally(uploaded_file, 'san-incentive', gcs_path)           

        if form.is_valid():
            deal = form.save(commit=False)
            if request.FILES.get('refDocs'):
                deal.refDocs.name = gcs_path
            deal.updated_by=request.session.get('mail_id')
            deal.save()
            return redirect('deal_list')
        else:
            print(form.errors)  # Optional: Debugging form errors
    else:
        form = DealForm(instance=deal)

    return render(request, 'owner/deal/deal_form.html', {
        'form': form,
        'action': 'Update',
        'title': 'Update Deal',
        'isCoordinator': isCoordinator,
        'deal': deal,
        'users': users,  # Pass users to populate dropdowns
    })

def deal_delete(request, pk):
    deal = get_object_or_404(Deal, pk=pk)
    if request.method == 'POST':
        deal.delete()
        return redirect('deal_list')
    return render(request, 'owner/deal/deal_confirm_delete.html', {'deal': deal})

# ---------- Annual Target Views ----------
def target_list(request):
    user_id = request.session.get('user_id')
    
    if not user_id:
        return []  # Or redirect to login
    try:
        current_user = UserProfile.objects.get(id=user_id)
    except UserProfile.DoesNotExist:
        return []
   
    search_query = request.GET.get('search', '').strip()
    start_year = request.GET.get('start_year', '').strip()
    end_year = request.GET.get('end_year', '').strip()

    allowed_roles = ['superadmin', 'accounts', 'admin']
    role_name = current_user.user_type.name.lower() if current_user.user_type else ''

    if role_name in allowed_roles:
    # Base queryset with related employee to avoid N+1
        targets = AnnualTarget.objects.all().select_related('employee').order_by('employee__fullname', 'financial_year')
    else:
        targets = AnnualTarget.objects.filter(
            Q(employee=current_user) | Q(employee__team_head=current_user)).select_related('employee').order_by('employee__fullname', 'financial_year')

    # Apply search filter on employee fullname or financial_year
    if search_query:
        targets = targets.filter(
            Q(employee__fullname__icontains=search_query) |
            Q(financial_year__icontains=search_query)
        )

    # Convert queryset to list for manual duplicate filtering
    targets_list = list(targets)

    # Remove duplicates based on (employee.id, financial_year)
    seen = set()
    unique_targets = []
    for t in targets_list:
        key = (t.employee.id, t.financial_year)
        if key not in seen:
            unique_targets.append(t)
            seen.add(key)

    # Filter by financial year range if both start_year and end_year are valid integers
    if start_year.isdigit() and end_year.isdigit():
        start_year_int = int(start_year)
        end_year_int = int(end_year)
        filtered_targets = []
        for t in unique_targets:
            try:
                fy_start = int(t.financial_year[:4])  # Assumes format "2023-24"
                if start_year_int <= fy_start <= end_year_int:
                    filtered_targets.append(t)
            except Exception:
                # Skip invalid formats
                pass
        unique_targets = filtered_targets

    # Paginate the results
    paginator = Paginator(unique_targets, 10)  # Show 10 targets per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'start_year': start_year,
        'end_year': end_year,
    }
    return render(request, 'owner/annual_target/target_list.html', context)

def target_export_excel(request):
    targets = AnnualTarget.objects.all().values(
        'employee__fullname', 'financial_year', 'annual_target_amount', 'net_salary'
    )
    df = pd.DataFrame(targets)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="annual_targets.xlsx"'
    df.to_excel(response, index=False)
    return response

def target_export_pdf(request):
    targets = AnnualTarget.objects.all().order_by('employee__fullname', 'financial_year')
    template = get_template('owner/annual_target/target_pdf.html')  # create this
    html = template.render({'targets': targets})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="annual_targets.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response
    
class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return str(obj)  # Or use float(obj) to convert to float instead of string
        return super().default(obj)

def target_create(request):
    users = UserProfile.objects.filter(user_type__name__in=['salesperson', 'saleshead'])
    if request.method == 'POST':
        form = AnnualTargetForm(request.POST)
        if form.is_valid():
            # Log the form data before saving
            print("Form data:", form.cleaned_data)
            annual_target = form.save()
            print(f"Annual target saved: {annual_target}")  # Log the saved instance
            return redirect('target_list')
        else:
            print("Form errors:", form.errors)  # Log form errors
            return render(request, 'owner/annual_target/target_form.html', {
                'form': form,
                'users': users,
                'financial_years': generate_financial_years(),
                'title': 'Create Annual Target',
                'action': 'Create'
            })
    else:
        form = AnnualTargetForm()

    return render(request, 'owner/annual_target/target_form.html', {
        'form': form,
        'users': users,
        'financial_years': generate_financial_years(),
        'title': 'Create Annual Target',
        'action': 'Create'
    })

def target_update(request, pk):
    target = get_object_or_404(AnnualTarget, pk=pk)
    users = UserProfile.objects.filter(user_type__name__in=['salesperson', 'saleshead'])
    if request.method == 'POST':
        form = AnnualTargetForm(request.POST, instance=target)
        if form.is_valid():
            # Save the form if it's valid
            form.save()
            return redirect('target_list')  # Redirect after saving
    else:
        form = AnnualTargetForm(instance=target)

    return render(request, 'owner/annual_target/target_form.html', {
        'form': form,
        'action': 'Update',
        'target': target,
        'users': users,
        'financial_years': generate_financial_years(),
        'title': 'Update Annual Target'
    })
def target_delete(request, pk):
    target = get_object_or_404(AnnualTarget, pk=pk)
    if request.method == 'POST':
        target.delete()
        return redirect('target_list')
    return render(request, 'owner/annual_target/target_confirm_delete.html', {'target': target})

def generate_financial_years():
    current_year = datetime.now().year
    return [f"{year}" for year in range(current_year -1, current_year + 4)]


def module(request):
    modules = Module.objects.all()

    edit_id = request.GET.get('edit')
    if edit_id:
        instance = get_object_or_404(Module, pk=edit_id)
    else:
        instance = None

    if request.method == 'POST':
        # Check for delete
        if 'delete_id' in request.POST:
            delete_id = request.POST.get('delete_id')
            Module.objects.filter(pk=delete_id).delete()
            return redirect('module')  # After delete, clean reload

        # Else, handle save
        form = ModuleForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            return redirect('module')  # After save, clean reload
    else:
        form = ModuleForm(instance=instance)

    return render(request, 'owner/settings/module.html', {
        'modules': modules,  # <<< correct name (small m)
        'form': form,
        'editing': bool(edit_id),
    })

def segment(request):
    edit_id = request.GET.get('edit')
    delete_id = request.POST.get('delete_id')

    # Handle Delete
    if delete_id:
        segment = get_object_or_404(Segment, id=delete_id)
        segment.delete()
        messages.success(request, "Segment deleted successfully.")
        return redirect('segment')  # Update with your url name

    # Handle Create or Edit
    if edit_id:
        segment = get_object_or_404(Segment, id=edit_id)
    else:
        segment = Segment()

    if request.method == 'POST' and not delete_id:
        form = SegmentForm(request.POST, instance=segment)
        if form.is_valid():
            form.save()
            messages.success(request, "Segment saved successfully.")
            return redirect('segment')  # Update with your url name
    else:
        form = SegmentForm(instance=segment)

    segments = Segment.objects.all()

    return render(request, 'owner/master/segment.html', {
        'form': form,
        'Segments': segments,
        'title': 'Add/Edit Segment',
        'button_label': 'Save Segment'
    })

from django.shortcuts import render, redirect, get_object_or_404
from .models import LeadSource
from .forms import LeadSourceForm

def leadsource(request):
    edit_id = request.GET.get('edit')  # fetch edit id from query params

    if request.method == 'POST':
        if 'delete_id' in request.POST:
            # Handle Delete
            delete_id = request.POST.get('delete_id')
            leadsource = get_object_or_404(LeadSource, pk=delete_id)
            leadsource.delete()
            return redirect('leadsource')  # Make sure your URL name is correct
        else:
            # Handle Create or Update
            leadsource_id = request.POST.get('id')
            instance = LeadSource.objects.filter(pk=leadsource_id).first() if leadsource_id else None
            form = LeadSourceForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                return redirect('leadsource')
    else:
        # GET Request
        instance = LeadSource.objects.filter(pk=edit_id).first() if edit_id else None
        form = LeadSourceForm(instance=instance)

    lead_sources = LeadSource.objects.all()

    return render(request, 'owner/master/leadsource.html', {
        'form': form,
        'lead_sources': lead_sources,  # <- fixed from 'roles' to 'lead_sources'
        'title': 'Manage Lead Source',
        'button_label': 'Update' if edit_id else 'Save',
    })

def roles(request):
    edit_id = request.GET.get('edit')  # fetch edit id from query params if any

    if request.method == 'POST':
        if 'delete_id' in request.POST:  # Delete operation
            delete_id = request.POST.get('delete_id')
            role = get_object_or_404(Role, pk=delete_id)
            role.delete()
            return redirect('roles')
        else:  # Create or Update operation
            role_id = request.POST.get('id')
            instance = Role.objects.filter(pk=role_id).first() if role_id else None
            form = RoleForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                return redirect('roles')
    else:
        instance = Role.objects.filter(pk=edit_id).first() if edit_id else None
        form = RoleForm(instance=instance)

    roles = Role.objects.all()
    return render(request, 'owner/settings/roles.html', {
        'form': form,
        'roles': roles,
        'title': 'Manage Roles',
        'button_label': 'Update' if edit_id else 'Save',
    })

def permission(request):
    roles = Role.objects.all()
    modules = Module.objects.all()

    if request.method == 'POST':
        for role in roles:
            for module in modules:
                can_add = bool(request.POST.get(f'permissions_{role.id}_{module.id}_add'))
                can_edit = bool(request.POST.get(f'permissions_{role.id}_{module.id}_edit'))
                can_delete = bool(request.POST.get(f'permissions_{role.id}_{module.id}_delete'))
                can_view = bool(request.POST.get(f'permissions_{role.id}_{module.id}_view'))

                # Always update_or_create regardless of True/False
                Permission.objects.update_or_create(
                    role=role,
                    module=module,
                    defaults={
                        'can_add': can_add,
                        'can_edit': can_edit,
                        'can_delete': can_delete,
                        'can_view': can_view,
                    }
                )

        return redirect('permission')

    # GET request
    permissions = Permission.objects.all()
    permission_matrix = {}

    for permission in permissions:
        permission_matrix.setdefault(permission.role_id, {})[permission.module_id] = {
            'can_add': permission.can_add,
            'can_edit': permission.can_edit,
            'can_delete': permission.can_delete,
            'can_view': permission.can_view,
        }

    return render(request, 'owner/settings/permissions.html', {
        'roles': roles,
        'modules': modules,
        'permission_matrix': permission_matrix,
    })

from django.shortcuts import render, redirect
from datetime import datetime
from .forms import IncentiveSetupForm
from .models import IncentiveSetup, SetupChargeSlab, TopperMonthSlab, HighValueDealSlab, Segment

def incentive_setup_create(request):
    current_year = datetime.now().year
    financial_years = [f"{year}" for year in range(current_year -1, current_year + 4)]
    months = [month for month in range(1, 13)]
    segments = Segment.objects.all()

    if request.method == 'POST':
        form = IncentiveSetupForm(request.POST)
        if form.is_valid():
            monthly_incentive = form.save()

            # ✅ Save SetupChargeSlabs with deal_type_setup
            deal_type_setup = request.POST.getlist('deal_type_setup[]')
            min_amounts = request.POST.getlist('setup_min_amount[]')
            max_amounts = request.POST.getlist('setup_max_amount[]')
            incentive_percentages = request.POST.getlist('setup_incentive_percentage[]')

            for deal_type, min_amt, max_amt, inc_perc in zip(deal_type_setup, min_amounts, max_amounts, incentive_percentages):
                SetupChargeSlab.objects.create(
                    incentive_setup=monthly_incentive,
                    deal_type_setup=deal_type,  # <-- This was missing
                    min_amount=min_amt or 0,
                    max_amount=max_amt or 0,
                    incentive_percentage=inc_perc or 0,
                    created_by=request.session.get('mail_id')
                )

            # ✅ Save TopperMonthSlabs
            deal_type_top = request.POST.getlist('deal_type_top[]')
            topper_segments = request.POST.getlist('segment[]')
            min_subs = request.POST.getlist('min_subscription[]')
            incentives = request.POST.getlist('incentive_percentage[]')

            for deal_type, segment_id, min_sub, inc_perc in zip(deal_type_top,topper_segments, min_subs, incentives):
                if segment_id:
                    segment_instance = get_object_or_404(Segment, pk=segment_id)
                    TopperMonthSlab.objects.create(
                        incentive_setup=monthly_incentive,
                        deal_type_top=deal_type,
                        segment=segment_instance,
                        min_subscription=min_sub or 0,
                        incentive_percentage=inc_perc or 0,
                        created_by=request.session.get('mail_id')
                    )

            # ✅ Save HighValueDealSlabs
            deal_types_high = request.POST.getlist('deal_type_high[]')
            min_values = request.POST.getlist('high_value_min_amount[]')
            max_values = request.POST.getlist('high_value_max_amount[]')
            high_value_incentives = request.POST.getlist('high_value_incentive_percentage[]')

            for deal_type, min_val, max_val, inc_perc in zip(deal_types_high, min_values, max_values, high_value_incentives):
                HighValueDealSlab.objects.create(
                    incentive_setup=monthly_incentive,
                    deal_type_high=deal_type,
                    min_amount=min_val or 0,
                    max_amount=max_val or 0,
                    incentive_percentage=inc_perc or 0,
                    created_by=request.session.get('mail_id')
                )

            return redirect('incentive_setup_list')
        else:
            print("Form errors:", form.errors)
    else:
        form = IncentiveSetupForm()

    return render(request, 'owner/incentive_setup/incentive_setup_form.html', {
        'form': form,
        'financial_years': financial_years,
        'title': 'Create Incentive Setup',
        'months': months,
        'segments': segments,
    })

from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils.dateparse import parse_date
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
import openpyxl
from .models import IncentiveSetup

def incentive_setup_list(request):
    search = request.GET.get('search', '')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    export = request.GET.get('export')

    incentives = IncentiveSetup.objects.all().order_by('-financial_year')

    # Search filter
    if search:
        incentives = incentives.filter(Q(financial_year__icontains=search))

    # Date range filter (assuming 'created_at' field exists)
    if from_date and to_date:
        from_date_parsed = parse_date(from_date)
        to_date_parsed = parse_date(to_date)
        if from_date_parsed and to_date_parsed:
            incentives = incentives.filter(created_at__range=(from_date_parsed, to_date_parsed))

    # Export options
    if export == 'xlsx':
        return export_incentives_to_xlsx(incentives)
    elif export == 'pdf':
        return export_incentives_to_pdf(incentives)

    # Pagination
    paginator = Paginator(incentives, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'owner/incentive_setup/incentive_setup_list.html', {
        'incentives': page_obj,
        'search': search,
        'from_date': from_date,
        'to_date': to_date,
    })

# XLSX Export
def export_incentives_to_xlsx(incentives):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Incentives"
    ws.append(['S.No.', 'Financial Year'])

    for i, incentive in enumerate(incentives, start=1):
        ws.append([i, incentive.financial_year])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="incentives.xlsx"'
    wb.save(response)
    return response

# PDF Export
def export_incentives_to_pdf(incentives):
    html = render_to_string('owner/incentive_setup/incentive_pdf.html', {'incentives': incentives})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="incentives.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response

def incentive_setup_delete(request, pk):
    # Get the IncentiveSetup or 404
    incentive = get_object_or_404(IncentiveSetup, pk=pk)
    
    if request.method == "POST":
        financial_year_to_delete = incentive.financial_year
        # Delete the incentive setup
        approved_deals_exist = Deal.objects.filter(
            dealWonDate__year=financial_year_to_delete,
            status='Approved' # <--- Adjust 'status' and 'Approved' to your Deal model's actual fields
        ).exists()

        if approved_deals_exist:
            messages.error(request, f"Cannot delete incentive setup for {financial_year_to_delete} as approved deals exist for this financial year.")
            request.session['show_modal'] = True
            return redirect('incentive_setup_list')
        else:
            # Delete the incentive setup
            incentive.delete()
            messages.success(request, f"Incentive setup for {financial_year_to_delete} deleted successfully.")
            return redirect('incentive_setup_list')

    # Confirmation page to delete
    return render(request, 'owner/incentive_setup/incentive_setup_confirm_delete.html', {'incentive': incentive})

def incentive_setup_view(request, pk):
    incentive = get_object_or_404(IncentiveSetup, pk=pk)

    current_year = datetime.now().year
    financial_years = [f"{year}" for year in range(current_year - 1, current_year + 4)]
    segments = Segment.objects.all()
    months = list(range(1, 13))

    setup_slabs = SetupChargeSlab.objects.filter(incentive_setup=incentive)
    topper_slabs = TopperMonthSlab.objects.filter(incentive_setup=incentive)
    highvalue_slabs = HighValueDealSlab.objects.filter(incentive_setup=incentive)

    return render(request, 'owner/incentive_setup/incentive_setup_view.html', {
        'incentive': incentive,
        'setup_slabs': setup_slabs,
        'topper_slabs': topper_slabs,
        'highvalue_slabs': highvalue_slabs,
        'segments': segments,
        'financial_years': financial_years,
        'months': months,
        'title': 'View Incentive Setup',
    })

def incentive_setup_update(request, pk):
    incentive = get_object_or_404(IncentiveSetup, pk=pk)

    current_year = datetime.now().year
    financial_years = [f"{year}" for year in range(current_year -1, current_year + 4)]
    segments = Segment.objects.all()
    months = list(range(1, 13))
    
    # --- Restriction: Prevent update if approved deals exist ---
    approved_deals_exist = Deal.objects.filter(
        dealWonDate__year=incentive.financial_year,
        status='Approved'
    ).exists()

    if approved_deals_exist:
        messages.error(request, f"Incentive setup for {incentive.financial_year} cannot be updated because approved deals already exist.")
        request.session['show_modal'] = True  # trigger modal in template
        return redirect('incentive_setup_list')

    if request.method == 'POST':
        form = IncentiveSetupForm(request.POST, instance=incentive)

        if form.is_valid():
            form.save()

            # --- Setup Slabs ---
            ids = request.POST.getlist('setup_id[]')
            deal_type_setup = request.POST.getlist('deal_type_setup[]')
            min_amounts = request.POST.getlist('setup_min_amount[]')
            max_amounts = request.POST.getlist('setup_max_amount[]')
            setup_incentive_percentage = request.POST.getlist('setup_incentive_percentage[]')

            if not validate_form_data_length(deal_type_setup, ids, min_amounts, max_amounts, setup_incentive_percentage):
                messages.error(request, "Setup slab data is incomplete or incorrectly structured.")
                return redirect('incentive_setup_update', pk=pk)

            for i in range(len(deal_type_setup)):
                try:
                    sid = ids[i]
                    slab = SetupChargeSlab.objects.get(id=sid) if sid else SetupChargeSlab(incentive_setup=incentive)
                    slab.deal_type_setup = deal_type_setup[i]
                    slab.min_amount = min_amounts[i]
                    slab.max_amount = max_amounts[i]
                    slab.incentive_percentage = setup_incentive_percentage[i]
                    slab.updated_by=request.session.get('mail_id')
                    slab.save()
                except Exception as e:
                    logger.error(f"[Setup Slab] Error at index {i}: {e}")
                    messages.error(request, f"Error saving Setup Slab #{i+1}: {e}")
                    return redirect('incentive_setup_update', pk=pk)

            # --- Topper Month Slabs ---
            ids = request.POST.getlist('topper_id[]')
            deal_type_top = request.POST.getlist('deal_type_top[]')
            segment_ids = request.POST.getlist('segment[]')
            min_subscription = request.POST.getlist('min_subscription[]')
            incentive_percentage = request.POST.getlist('incentive_percentage[]')
            print("hiii1")
            if not validate_form_data_length(deal_type_top, ids, segment_ids, min_subscription, incentive_percentage):
                messages.error(request, "Topper slab data is incomplete or incorrectly structured.")
                print(request, "Topper slab data is incomplete or incorrectly structured.")
                return redirect('incentive_setup_update', pk=pk)

            for i in range(len(deal_type_top)):
                try:
                    print(f"Processing deal {i+1} / {len(deal_type_top)}")
                    tid = ids[i]
                    print(f"tid: {tid}, deal_type_top: {deal_type_top[i]}, segment_id: {segment_ids[i]}")

                    slab = TopperMonthSlab.objects.get(id=tid) if tid else TopperMonthSlab(incentive_setup=incentive)

                    try:
                        segment_instance = Segment.objects.get(id=segment_ids[i])
                        print(f"Found segment: {segment_instance}")
                    except Segment.DoesNotExist:
                        error_msg = f"Segment with ID {segment_ids[i]} not found."
                        logger.error(f"[Topper Slab] {error_msg}")
                        messages.error(request, error_msg)
                        return redirect('incentive_setup_update', pk=pk)

                    slab.deal_type_top = deal_type_top[i]
                    slab.segment = segment_instance
                    slab.min_subscription = min_subscription[i]
                    slab.incentive_percentage = incentive_percentage[i]
                    slab.updated_by=request.session.get('mail_id')
                    slab.save()
                    print(f"Saved slab: {slab}")
                except Exception as e:
                    logger.error(f"[Topper Slab] Error at index {i}: {e}")
                    messages.error(request, f"Error saving Topper Slab #{i+1}: {e}")
                    return redirect('incentive_setup_update', pk=pk)

            # --- High Value Deal Slabs ---
            ids = request.POST.getlist('highvalue_id[]')
            deal_type_high = request.POST.getlist('deal_type_high[]')
            high_value_min_amount = request.POST.getlist('high_value_min_amount[]')
            high_value_max_amount = request.POST.getlist('high_value_max_amount[]')
            high_value_incentive_percentage = request.POST.getlist('high_value_incentive_percentage[]')

            if not validate_form_data_length(deal_type_high, ids, high_value_min_amount, high_value_max_amount, high_value_incentive_percentage):
                messages.error(request, "High value slab data is incomplete or incorrectly structured.")
                return redirect('incentive_setup_update', pk=pk)

            for i in range(len(deal_type_high)):
                try:
                    hid = ids[i]
                    slab = HighValueDealSlab.objects.get(id=hid) if hid else HighValueDealSlab(incentive_setup=incentive)
                    slab.deal_type_high = deal_type_high[i]
                    slab.min_amount = high_value_min_amount[i]
                    slab.max_amount = high_value_max_amount[i]
                    slab.incentive_percentage = high_value_incentive_percentage[i]
                    slab.updated_by=request.session.get('mail_id')
                    slab.save()
                except Exception as e:
                    logger.error(f"[High Value Slab] Error at index {i}: {e}")
                    messages.error(request, f"Error saving High Value Slab #{i+1}: {e}")
                    return redirect('incentive_setup_update', pk=pk)

            messages.success(request, "Incentive Setup updated successfully.")
            return redirect('incentive_setup_list')
        else:
            messages.error(request, "Please fix the errors in the form.")
    else:
        form = IncentiveSetupForm(instance=incentive)

    setup_slabs = SetupChargeSlab.objects.filter(incentive_setup=incentive)
    topper_slabs = TopperMonthSlab.objects.filter(incentive_setup=incentive)
    highvalue_slabs = HighValueDealSlab.objects.filter(incentive_setup=incentive)

    return render(request, 'owner/incentive_setup/incentive_setup_form.html', {
        'form': form,
        'setup_formset_slabs': setup_slabs,
        'topper_month_slabs': topper_slabs,
        'high_value_slabs': highvalue_slabs,
        'segments': segments,
        'financial_years': financial_years,
        'months': months,
        'title': 'Update Incentive Setup',
    })


def deal_approve(request, pk):
    deal = get_object_or_404(Deal, pk=pk)
    if deal.status != 'Approved':
        deal.status = 'Approved'
        deal.updated_by = request.user
        deal.save()

        # Call Rule Engine after approval
        # DealRuleEngine(deal).run_rules()  # Encapsulate logic inside a method        
    return redirect('deal_list')


def salesteam(request):
    sales_heads = UserProfile.objects.filter(user_type__name__in=['saleshead'])
    salespersons = UserProfile.objects.filter(user_type__name__in=['salesperson'])

    return render(request, 'owner/master/SalesTeamHierarchy.html', {
        'sales_heads': sales_heads,
        'salespersons': salespersons,
    })


from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from .models import PayoutTransaction
import openpyxl
from xhtml2pdf import pisa
from django.template.loader import render_to_string

def payout(request):
    search_query = request.GET.get('search', '')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    export_format = request.GET.get('export')
    order_by = request.GET.get('order_by', '-created_at')  # Default to descending creation date

    payouts = PayoutTransaction.objects.all()

    # Apply search
    if search_query:
        payouts = payouts.filter(
            Q(user__fullname__icontains=search_query) |
            Q(deal__clientName__icontains=search_query)
        )

    # Apply date filter
    if from_date and to_date:
        from_date = parse_date(from_date)
        to_date = parse_date(to_date)
        if from_date and to_date:
            to_date = datetime.combine(to_date, datetime.max.time())
            from_date = datetime.combine(from_date, datetime.min.time())
            payouts = payouts.filter(created_at__range=(from_date, to_date))

    # Apply ordering
    try:
        payouts = payouts.order_by(order_by)
    except:
        payouts = payouts.order_by('-created_at')  # Fallback

    # Export
    if export_format == 'xlsx':
        return export_payouts_to_xlsx(payouts)
    elif export_format == 'pdf':
        return export_payouts_to_pdf(payouts)

    # Pagination
    paginator = Paginator(payouts, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'owner/payout/payout_list.html', {
        'payouts': page_obj,
        'search_query': search_query,
        'from_date': from_date,
        'to_date': to_date,
        'order_by': order_by
    })

def export_payouts_to_xlsx(payouts):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payouts"
    ws.append(['User', 'Deal ID', 'Person Type', 'Status', 'Amount'])

    for p in payouts:
        ws.append([
            p.user.fullname,
            p.deal_id,
            p.incentive_person_type,
            p.payout_status,
            p.payout_amount
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="payouts.xlsx"'
    wb.save(response)
    return response

def export_payouts_to_pdf(payouts):
    html = render_to_string('owner/payout/pdf_template.html', {'payouts': payouts})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="payouts.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response



def transaction(request):
    search = request.GET.get('search', '')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    export = request.GET.get('export')

    transactions = Transaction.objects.all().order_by('-transaction_date')

    # Filter by search
    if search:
        transactions = transactions.filter(
            Q(user__fullname__icontains=search) |
            Q(deal_id__clientName__icontains=search) |
            Q(transaction_type__icontains=search) |
            Q(incentive_component_type__icontains=search) |
            Q(eligibility_message__icontains=search) |
            Q(notes__icontains=search)
        )

    # Filter by date
    if from_date and to_date:
        from_parsed = parse_date(from_date)
        to_parsed = parse_date(to_date)
        if from_parsed and to_parsed:
            transactions = transactions.filter(transaction_date__range=(from_parsed, to_parsed))

    # Export
    if export == 'xlsx':
        return export_transaction_xlsx(transactions)
    elif export == 'pdf':
        return export_transaction_pdf(transactions)

    # Pagination
    paginator = Paginator(transactions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'owner/reports/transaction.html', {
        'page_obj': page_obj,
        'search': search,
        'from_date': from_date,
        'to_date': to_date,
    })

def export_transaction_xlsx(transactions):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(['Deal ID', 'Version', 'Type', 'Component', 'Amount', 'Frozen', 'Latest', 'Eligibility', 'Message', 'Date', 'Notes'])

    for txn in transactions:
        ws.append([
            txn.deal_id, txn.version, txn.transaction_type, txn.incentive_component_type,
            txn.amount, txn.freeze, txn.is_latest, txn.eligibility_status,
            txn.eligibility_message, txn.transaction_date.strftime('%Y-%m-%d'),
            txn.notes or '-'
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="transactions.xlsx"'
    wb.save(response)
    return response

def export_transaction_pdf(transactions):
    html = render_to_string('owner/reports/transaction_pdf.html', {'transactions': transactions})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="transactions.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response

def transaction_export_excel(request):
    search = request.GET.get('search', '')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    transactions = Transaction.objects.all()

    if search:
        transactions = transactions.filter(
            Q(deal_id__clientName__icontains=search) |
            Q(notes__icontains=search) |
            Q(transaction_type__icontains=search)
        )
 
    # Date filter   
    if from_date and from_date != 'None':
        parsed_from_date = parse_date(from_date)
        if parsed_from_date:
            transactions = transactions.filter(transaction_date__gte=parsed_from_date)

    if to_date and to_date != 'None':
        parsed_to_date = parse_date(to_date)
        if parsed_to_date:
            transactions = transactions.filter(transaction_date__lte=parsed_to_date)

    data = tablib.Dataset()
    data.headers = ['Deal Name', 'Version', 'Type', 'Component', 'Amount', 'Frozen', 'Latest', 'Eligibility', 'Message', 'Date', 'Notes']

    for txn in transactions:
        data.append([
            txn.deal_id,
            txn.version,
            txn.transaction_type,
            txn.incentive_component_type,
            txn.amount,
            txn.freeze,
            txn.is_latest,
            txn.eligibility_status,
            txn.eligibility_message,
            txn.transaction_date.replace(tzinfo=None) if txn.transaction_date else None,
            txn.notes or '-'
        ])

    response = HttpResponse(data.xlsx, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="transactions.xlsx"'
    return response

def transaction_export_pdf(request):
    search = request.GET.get('search', '')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    transactions = Transaction.objects.all()
    if search:
        transactions = transactions.filter(
            Q(user__fullname__icontains=search) |
            Q(deal_id__clientName__icontains=search) |
            Q(notes__icontains=search) |
            Q(transaction_type__icontains=search)
        )
    # Date filter   
    if from_date and from_date != 'None':
        parsed_from_date = parse_date(from_date)
        if parsed_from_date:
            transactions = transactions.filter(transaction_date__gte=parsed_from_date)

    if to_date and to_date != 'None':
        parsed_to_date = parse_date(to_date)
        if parsed_to_date:
            transactions = transactions.filter(transaction_date__lte=parsed_to_date)

    template = get_template('owner/reports/transaction_pdf.html')
    html = template.render({'transactions': transactions})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="transactions.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response

def targettransaction(request):
    search = request.GET.get('search', '')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    transactions = TargetTransaction.objects.all()

    # Search filter (e.g., deal_id, message)
    if search:
        transactions = transactions.filter(
            Q(user__fullname__icontains=search) |
            Q(deal__clientName__icontains=search) |
            Q(notes__icontains=search) |
            Q(transaction_type__icontains=search)
        )

    # Date filter   
    if from_date and from_date != 'None':
        parsed_from_date = parse_date(from_date)
        if parsed_from_date:
            transactions = transactions.filter(transaction_date__gte=parsed_from_date)

    if to_date and to_date != 'None':
        parsed_to_date = parse_date(to_date)
        if parsed_to_date:
            transactions = transactions.filter(transaction_date__lte=parsed_to_date)    

    transactions = transactions.order_by('-transaction_date')

    paginator = Paginator(transactions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'owner/reports/targettransaction.html', {
        'page_obj': page_obj,
        'search': search,
        'from_date': from_date,
        'to_date': to_date,
    })
def targettransaction_export_excel(request):
    search = request.GET.get('search', '')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    transactions = TargetTransaction.objects.all()

    if search:
        transactions = transactions.filter(
            Q(user__fullname__icontains=search) |
            Q(deal__clientName__icontains=search) |
            Q(notes__icontains=search) |
            Q(transaction_type__icontains=search)
        )
 
    # Date filter   
    if from_date and from_date != 'None':
        parsed_from_date = parse_date(from_date)
        if parsed_from_date:
            transactions = transactions.filter(transaction_date__gte=parsed_from_date)

    if to_date and to_date != 'None':
        parsed_to_date = parse_date(to_date)
        if parsed_to_date:
            transactions = transactions.filter(transaction_date__lte=parsed_to_date)

    data = tablib.Dataset()
    data.headers = ['Deal Name', 'Version', 'Type', 'Component', 'Amount', 'Frozen', 'Latest', 'Eligibility', 'Message', 'Date', 'Notes']

    for txn in transactions:
        data.append([
            txn.deal,
            txn.version,
            txn.transaction_type,
            txn.incentive_component_type,
            txn.amount,
            txn.freeze,
            txn.is_latest,
            txn.eligibility_status,
            txn.eligibility_message,
            txn.transaction_date.replace(tzinfo=None) if txn.transaction_date else None,
            txn.notes or '-'
        ])

    response = HttpResponse(data.xlsx, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="targettransaction.xlsx"'
    return response

def targettransaction_export_pdf(request):
    search = request.GET.get('search', '')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    transactions = TargetTransaction.objects.all()
    if search:
        transactions = transactions.filter(
            Q(deal__clientName__icontains=search) |
            Q(notes__icontains=search) |
            Q(transaction_type__icontains=search)
        )
    # Date filter   
    if from_date and from_date != 'None':
        parsed_from_date = parse_date(from_date)
        if parsed_from_date:
            transactions = transactions.filter(transaction_date__gte=parsed_from_date)

    if to_date and to_date != 'None':
        parsed_to_date = parse_date(to_date)
        if parsed_to_date:
            transactions = transactions.filter(transaction_date__lte=parsed_to_date)

    template = get_template('owner/reports/transaction_pdf.html')
    html = template.render({'targettransaction': transactions})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="targettransaction.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response

def backup_db_view(request):
    if request.method == "POST":
        message = upload_db_to_gcs()
        return HttpResponse(message)
    return HttpResponse("Use POST to trigger backup.")

def schedulelog(request):
    months = range(1, 13)
    current_month = datetime.now().month 
    current_year = datetime.now().year
    years = [current_year - 1, current_year] 
    monthly_jobs = ScheduledJob.objects.filter(job_type='monthly').order_by('next_run')
    yearly_jobs = ScheduledJob.objects.filter(job_type='yearly').order_by('next_run')
    logs = JobRunLog.objects.all()  # optional, if you show logs
           
    return render(request, 'owner/activity/schedule.html', {
        'monthly_jobs': monthly_jobs,
        'yearly_jobs': yearly_jobs,
        "years": years,
        'months': months,  
        'current_month': current_month,       
        'logs': logs
    })

def changelog(request):
    changelogs = ChangeLog.objects.all().order_by('-created_at')[:50]  # Limit for performance
    paginator = Paginator(changelogs, 10)  # Show 10 logs per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'owner/activity/changelog.html', {'page_obj': page_obj})

def run_now(request, job, runmonth, runyear):
    run_month = runmonth or datetime.now().month
    run_year = runyear or datetime.now().year

    print(f"Running {job} job for month: {run_month}, year: {run_year}")

    if job == "monthly":
        monthly_sales_incentive(run_year, run_month)
    elif job == "annual":
        annual_target_achievement(run_year)

    return redirect('schedulelog')

def mark_payout_paid(request, payout_id):
    payout = get_object_or_404(PayoutTransaction, pk=payout_id)

    if payout.payout_status != 'ReadyToPay':
        return JsonResponse({'status': 'invalid status'}, status=400)

    payout.payout_status = 'Paid'
    payout.save(update_fields=['payout_status'])
    return JsonResponse({'status': 'success'})

def bulk_mark_paid(request):
    try:
        data = json.loads(request.body)        
        ids = data.get('payout_ids', [])
        if not isinstance(ids, list) or not ids:
            return JsonResponse({"error": "Invalid or missing payout_ids."}, status=400)
        
        ids = list(map(int, ids))
        queryset = PayoutTransaction.objects.filter(pk__in=ids, payout_status='ReadyToPay')            
        updated = queryset.update(payout_status='Paid')

        return JsonResponse({"success": True, "updated_count": updated})

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON."}, status=400)
    except Exception as e:
        print(f"Error: {e} .") 
        return JsonResponse({"error": str(e)}, status=500)

def change_password(request):
    from django.contrib.auth.hashers import check_password, make_password
    user_id = request.session.get('user_id')  # Assuming session stores user_id
    if not user_id:
        return redirect('login')

    user = UserProfile.objects.get(id=user_id)

    if request.method == 'POST':
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        if not check_password(current_password, user.password):
            return render(request, 'change_password.html', {'error': 'Incorrect current password'})

        if new_password != confirm_password:
            return render(request, 'change_password.html', {'error': 'New passwords do not match'})

        user.password = make_password(new_password)
        user.save()

        return render(request, 'change_password.html', {'success': 'Password changed successfully!'})

    return render(request, 'change_password.html')